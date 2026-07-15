#!/usr/bin/env python3
import sys
import questionary
from openpyxl import load_workbook
from src.styles import console, custom_style
from src.ui import banner
from src.excel_utils import get_row_color
from src.file_picker import pick_file, pick_sheet
from src.single_sheet import (
    view_data,
    keyword_search,
    export_clean_sheet,
    deduplicate_keywords,
)
from src.multi_sheet import (
    get_multi_sheet_data,
    view_multi_sheets,
    keyword_search_multi,
    export_clean_multi,
)
from src.instructions import show_instructions


def main():
    try:
        filepath = pick_file()
        wb = load_workbook(filepath)
        sheet_names = pick_sheet(wb)

        while True:
            banner()

            merged_data = get_multi_sheet_data(wb, sheet_names)
            total = sum(len(rows) for rows in merged_data.values())
            green_count = 0
            red_count = 0
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                for row_idx, _ in merged_data[sheet_name]:
                    row = [ws.cell(row=row_idx, column=1)]
                    if get_row_color(row) == "green":
                        green_count += 1
                    elif get_row_color(row) == "red":
                        red_count += 1

            sheet_display = (
                sheet_names[0]
                if len(sheet_names) == 1
                else f"{len(sheet_names)} sheets"
            )
            merge_indicator = "" if len(sheet_names) == 1 else " [MERGED]"

            console.print(
                f"  📂 [bright_cyan]{filepath.name}[/bright_cyan]  →  "
                f"📄 [bright_cyan]{sheet_display}[/bright_cyan]{merge_indicator}  |  "
                f"[white]{total}[/white] rows  "
                f"[green]{green_count} ✔[/green]  [red]{red_count} ✘[/red]  "
                f"[dim]{total - green_count - red_count} unmarked[/dim]\n"
            )

            action = questionary.select(
                "What would you like to do?",
                choices=[
                    "👁   View Data",
                    "🔍  Keyword Search",
                    "📤  Export Clean Sheet (remove irrelevant)",
                    "🧹  Remove Duplicate Keywords",
                    "📄  Switch Worksheet(s)",
                    "📂  Switch File",
                    "❓  Help & Instructions",
                    "🚪  Exit",
                ],
                style=custom_style,
                instruction="(↑/↓ to move, Enter to select)",
            ).ask()

            if action is None or "Exit" in action:
                banner()
                console.print("  [dim]Goodbye! 👋[/dim]\n")
                break

            elif "View" in action:
                if len(sheet_names) == 1:
                    view_data(wb[sheet_names[0]])
                else:
                    view_multi_sheets(wb, sheet_names)

            elif "Search" in action:
                if len(sheet_names) == 1:
                    keyword_search(wb[sheet_names[0]], wb, filepath)
                else:
                    keyword_search_multi(wb, sheet_names, filepath)

            elif "Export" in action:
                if len(sheet_names) == 1:
                    export_clean_sheet(wb[sheet_names[0]], wb, filepath)
                else:
                    export_clean_multi(wb, sheet_names, filepath)

            elif "Duplicate" in action:
                if len(sheet_names) == 1:
                    deduplicate_keywords(wb[sheet_names[0]], wb, filepath)
                else:
                    banner()
                    console.print(
                        "[yellow]  Deduplication is per-sheet. Please switch to a single sheet.[/yellow]\n"
                    )
                    questionary.press_any_key_to_continue(style=custom_style).ask()

            elif "Help" in action:
                show_instructions()

            elif "Switch Worksheet" in action:
                sheet_names = pick_sheet(wb)

            elif "Switch File" in action:
                filepath = pick_file()
                wb = load_workbook(filepath)
                sheet_names = pick_sheet(wb)

    except KeyboardInterrupt:
        console.print("\n  [dim]Interrupted. Goodbye! 👋[/dim]\n")
        sys.exit(0)


if __name__ == "__main__":
    main()
