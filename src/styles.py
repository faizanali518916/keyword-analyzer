import sys
from pathlib import Path

import questionary
from questionary import Style
from rich.console import Console
from openpyxl.styles import PatternFill

console = Console()

custom_style = Style(
    [
        ("qmark", "fg:#E91E63 bold"),
        ("question", "fg:#FFFFFF bold"),
        ("answer", "fg:#00BCD4 bold"),
        ("pointer", "fg:#E91E63 bold"),
        ("highlighted", "fg:#E91E63 bold"),
        ("selected", "fg:#00BCD4"),
        ("separator", "fg:#757575"),
        ("instruction", "fg:#9E9E9E"),
        ("text", "fg:#FFFFFF"),
    ]
)

RED_FILL = PatternFill(start_color="FF4C4C", end_color="FF4C4C", fill_type="solid")
GREEN_FILL = PatternFill(start_color="4CFF4C", end_color="4CFF4C", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
